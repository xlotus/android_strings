# Android strings python tools

import os.path
import shutil
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl import load_workbook

FILE_SEPARATOR = "/"
SPECIAL_SEPARATOR = "#"
POINT = "."
VALUES = "values"
XLSX_TAIL = ".xlsx"

DIR_PROJECT = "/Users/mac/works/examples/ShareitHello"
DIR_DEST = "./.temp_strings"
DIR_DEST_XLSX = "./.out_xlsx"


def get_temp_string_file_path(src_file_path):
    length = len(DIR_PROJECT)
    if not DIR_PROJECT.endswith(FILE_SEPARATOR):
        length = length + 1
    sub_path = src_file_path[length:]
    dest_file_name = sub_path.replace(FILE_SEPARATOR, SPECIAL_SEPARATOR)
    return os.path.join(DIR_DEST, dest_file_name)


# return xlsx file name and language
# app#src#main#res#values-en#strings.xml
def get_xlsx_file_info(temp_string_file_name):
    value_index = temp_string_file_name.index(VALUES)
    name_index = temp_string_file_name.rindex(SPECIAL_SEPARATOR)
    value_str = temp_string_file_name[value_index:name_index]
    value_end_index = name_index
    point_index = temp_string_file_name.rindex(POINT)
    if value_str.__contains__("-"):
        value_end_index = value_index + value_str.index("-")
    strings_file_name = temp_string_file_name[name_index + 1:point_index]
    xlsx_file_name = temp_string_file_name[0:value_index] + VALUES + SPECIAL_SEPARATOR + strings_file_name + XLSX_TAIL
    language = "default"
    if value_end_index < name_index:
        language = temp_string_file_name[value_end_index + 1:name_index]
    print(xlsx_file_name, language)
    return xlsx_file_name, language


def copy_file(file_path):
    if file_path.endswith("strings.xml"):
        dest_path = get_temp_string_file_path(file_path)
        print(dest_path)
        shutil.copy(file_path, dest_path)


def copy_strings(dir_src, dir_dest):
    file_list = os.listdir(dir_src)
    for file_name in file_list:
        full_file_path = os.path.join(dir_src, file_name)
        if os.path.isfile(full_file_path):
            copy_file(full_file_path)
        else:
            copy_strings(full_file_path, dir_dest)


def copy_strings_enter():
    if os.path.exists(DIR_DEST):
        shutil.rmtree(DIR_DEST)
    os.mkdir(DIR_DEST)
    copy_strings(DIR_PROJECT, DIR_DEST)


def get_text(element):
    value = element.text
    if value is None:
        value = ""
    if list(element) is not None:
        for child in list(element):
            child_text = ET.tostring(child, "utf-8").decode()
            if child_text is not None:
                value += child_text
    return value


def treat_src_files():
    print("treat_src_files")
    if os.path.exists(DIR_DEST_XLSX):
        shutil.rmtree(DIR_DEST_XLSX)
    os.mkdir(DIR_DEST_XLSX)
    file_list = os.listdir(DIR_DEST)
    for file_name in file_list:
        print(file_name)
        xlsx_file_name, language = get_xlsx_file_info(file_name)

        tree = ET.parse(os.path.join(DIR_DEST, file_name))
        string_list = tree.getroot()
        xlsx_file_path = os.path.join(DIR_DEST_XLSX, xlsx_file_name)
        if os.path.exists(xlsx_file_path):
            wb = load_workbook(xlsx_file_path)
        else:
            wb = Workbook()
            ws = wb.active
            wb.remove(ws)
        ws = wb.create_sheet(language)
        ws.append(["name", "translatable", "value"])
        for strElement in string_list:
            name = strElement.attrib["name"]
            value = get_text(strElement)

            # print(name + " : " + value)
            translatable = ""
            if strElement.keys().__contains__("translatable"):
                translatable = strElement.attrib["translatable"]
            ws.append([name, translatable, value])
        wb.save(xlsx_file_path)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    copy_strings_enter()
    treat_src_files()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
